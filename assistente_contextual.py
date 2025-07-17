import os
import json
import requests
from dotenv import load_dotenv

# Dependências para RAG e embeddings
from langchain_community.vectorstores import FAISS
from langchain_huggingface import HuggingFaceEmbeddings

# Novas dependências para manipulação de arquivos
import fitz  # PyMuPDF
import docx
import openpyxl

# --- SEÇÃO 1: FUNÇÃO DE COMUNICAÇÃO COM O GATEWAY ---

def chamar_servidor_gateway(endpoint: str, payload: dict) -> str:
    """Função centralizada para chamar endpoints do nosso servidor gateway."""
    try:
        url = f"http://127.0.0.1:8000/{endpoint.strip('/')}"
        response = requests.post(url, json=payload, timeout=300)
        response.raise_for_status()
        return response.json().get("texto_gerado", f"ERRO: Resposta inválida do endpoint /{endpoint}")
    except requests.exceptions.RequestException as e:
        return f"ERRO DE CONEXÃO com o Servidor Gateway: {e}"

# --- SEÇÃO 2: NOVAS FUNÇÕES PARA EXTRAÇÃO DE TEXTO ---

def extrair_texto_de_pdf(caminho_arquivo: str) -> str:
    try:
        with fitz.open(caminho_arquivo) as doc:
            texto = "".join(page.get_text() for page in doc)
        return texto
    except Exception as e:
        print(f"     ❌ ERRO ao ler PDF {caminho_arquivo}: {e}")
        return ""

def extrair_texto_de_docx(caminho_arquivo: str) -> str:
    try:
        doc = docx.Document(caminho_arquivo)
        return "\n".join([para.text for para in doc.paragraphs])
    except Exception as e:
        print(f"     ❌ ERRO ao ler DOCX {caminho_arquivo}: {e}")
        return ""

def extrair_texto_de_xlsx(caminho_arquivo: str) -> str:
    try:
        workbook = openpyxl.load_workbook(caminho_arquivo)
        texto_completo = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            texto_completo.append(f"--- Planilha: {sheet_name} ---\n")
            for row in sheet.iter_rows(values_only=True):
                texto_linha = "\t".join([str(cell) if cell is not None else "" for cell in row])
                texto_completo.append(texto_linha)
        return "\n".join(texto_completo)
    except Exception as e:
        print(f"     ❌ ERRO ao ler XLSX {caminho_arquivo}: {e}")
        return ""
        
def extrair_texto_de_fontes(fontes: list) -> str:
    """
    Lê uma lista de arquivos de diferentes formatos e retorna todo o conteúdo
    como uma única string de texto.
    """
    print("\n-> Extraindo texto de todas as fontes...")
    textos_completos = []
    
    for fonte in fontes:
        print(f"   - Processando: {fonte}")
        if not os.path.exists(fonte):
            print(f"     ⚠️ AVISO: Arquivo não encontrado, será ignorado: {fonte}")
            continue

        extensao = os.path.splitext(fonte)[1].lower()
        texto_extraido = ""

        if extensao == ".pdf":
            texto_extraido = extrair_texto_de_pdf(fonte)
        elif extensao in [".txt", ".json"]:
            try:
                with open(fonte, 'r', encoding='utf-8') as f:
                    texto_extraido = f.read()
            except Exception as e:
                print(f"     ❌ ERRO ao ler arquivo de texto {fonte}: {e}")
        elif extensao == ".docx":
            texto_extraido = extrair_texto_de_docx(fonte)
        elif extensao == ".xlsx":
            texto_extraido = extrair_texto_de_xlsx(fonte)
        else:
            print(f"     ⚠️ AVISO: Tipo de arquivo não suportado, será ignorado: {extensao}")
            continue
        
        if texto_extraido:
            header = f"\n\n--- INÍCIO DO DOCUMENTO: {os.path.basename(fonte)} ---\n\n"
            footer = f"\n\n--- FIM DO DOCUMENTO: {os.path.basename(fonte)} ---\n\n"
            textos_completos.append(header + texto_extraido + footer)

    if not textos_completos:
        print("   ❌ ERRO: Nenhum texto pôde ser extraído dos arquivos fornecidos.")
        return ""
        
    print("✅ Extração de texto concluída.")
    return "".join(textos_completos)


# --- SEÇÃO 3: LÓGICA DE CHAT ---

def loop_chat_rag(db: FAISS, nome_especialista: str):
    # ... (lógica do RAG permanece idêntica) ...
    print(f"\n✅ Especialista '{nome_especialista}' pronto! (Comunicação via Servidor Gateway)")
    usar_resumo = input("Deseja SUMARIZAR o contexto antes de enviar? (s/n, padrão 'n'): ").lower() == 's'
    print("   Digite 'sair' a qualquer momento para terminar.")
    
    while True:
        pergunta = input(f"\n🤖 Você pergunta para '{nome_especialista}': ")
        if pergunta.strip().lower() == 'sair': break
        print("   -> Fase 1: Buscando documentos...")
        docs_relevantes = db.similarity_search(pergunta, k=15)
        if not docs_relevantes: print("\n💡 Resposta: Não encontrei documentos relevantes."); continue
        contexto_original = "\n\n".join([doc.page_content for doc in docs_relevantes])
        contexto_para_geracao = contexto_original
        
        if usar_resumo:
            print("   -> Fase 2a: Solicitando sumarização ao servidor...")
            payload = {"contexto": contexto_original, "pergunta": pergunta}
            contexto_para_geracao = chamar_servidor_gateway("sumarizar", payload)
        
        print("   -> Fase 2b: Solicitando geração RAG ao servidor...")
        payload = {"contexto": contexto_para_geracao, "pergunta": pergunta}
        resposta_final = chamar_servidor_gateway("gerar_rag", payload)
        
        print("\n💡 Resposta do Especialista (via Servidor Gateway):")
        print(resposta_final)


def loop_chat_puro():
    # ... (lógica do chat puro permanece idêntica) ...
    print(f"\n✅ Chat direto com o modelo principal ativo no servidor: '{nome_modelo_principal_ativo}'")
    while True:
        pergunta = input("\n🤖 Você pergunta: ")
        if pergunta.strip().lower() == 'sair': break
        print("   ...pensando (via servidor gateway)...")
        response = chamar_servidor_gateway("gerar", {"prompt": pergunta})
        print("\n💡 Resposta do Gateway:")
        print(response)

# ATUALIZADO: O loop de análise de arquivos agora se parece muito com o loop de RAG
def loop_analise_de_arquivos():
    fontes = CONTEXTOS_DISPONIVEIS.get("pdf_openrouter", {}).get("fontes", [])
    if not fontes: print("ERRO: Nenhuma fonte definida para 'pdf_openrouter' em contexts.json."); return
    
    # Extrai todo o texto dos arquivos para uma única variável
    contexto_dos_arquivos = extrair_texto_de_fontes(fontes)
    
    if not contexto_dos_arquivos:
        return # Encerra se nenhum texto foi extraído
        
    print("\n✅ Documentos processados e prontos para análise.")
    while True:
        pergunta = input("\n🤖 Você pergunta sobre os documentos: ")
        if pergunta.lower() == 'sair': break
            
        print("   ...solicitando análise ao servidor gateway (via RAG)...")
        # Reutiliza o endpoint /gerar_rag, enviando todo o texto como contexto
        payload = {"contexto": contexto_dos_arquivos, "pergunta": pergunta}
        resposta = chamar_servidor_gateway("gerar_rag", payload)
        
        print("\n💡 Resposta do Assistente de Documentos:")
        print(resposta)


# --- EXECUÇÃO PRINCIPAL ---

if __name__ == "__main__":
    print("-> Configurando o ambiente do assistente (Controlador Principal)...")
    try:
        with open("contexts.json", 'r', encoding='utf-8') as f:
            CONTEXTOS_DISPONIVEIS = json.load(f)
        with open("prompts.json", 'r', encoding='utf-8') as f:
            PROMPTS_CONFIG = json.load(f)
        with open("config_modelo_local.json", 'r', encoding='utf-8') as f:
            CONFIG = json.load(f)
        serv_princ_cfg = CONFIG.get("servicos", {}).get("gerador_principal", {})
        nome_modelo_principal_ativo = os.path.basename(serv_princ_cfg.get('path_gguf')) if serv_princ_cfg.get('tipo') == 'local' else serv_princ_cfg.get('id_openrouter')
    except Exception as e:
        print(f"AVISO: Não foi possível ler todas as configurações: {e}")

    embeddings = HuggingFaceEmbeddings(model_name="all-MiniLM-L6-v2")
    PASTA_BASE_INDICES = "indices_rag"
    print("✅ Ambiente do cliente configurado.")

    print("\n--- Assistente de IA com Servidor Gateway ---")
    print("Escolha o modo de operação:")
    print(f"  1. Conversa Geral (com o modelo principal do servidor: {nome_modelo_principal_ativo})")
    print("  2. Analisar Documentos (via Servidor Gateway)") # Menu atualizado
    
    opcoes_rag = {}
    i = 3
    for id_ctx, definicao_ctx in CONTEXTOS_DISPONIVEIS.items():
        if id_ctx == "pdf_openrouter": continue # Não mostra a opção de PDF no menu RAG
        status = "✅ Indexado" if os.path.exists(os.path.join(PASTA_BASE_INDICES, id_ctx)) else "❌ Não Indexado"
        opcoes_rag[str(i)] = {"id": id_ctx, "nome": definicao_ctx['nome_exibicao'], "status": status}
        print(f"  {i}. Especialista RAG: {definicao_ctx['nome_exibicao']} ({status})")
        i += 1

    escolha_principal = input("\nDigite o número da sua opção: ")

    if escolha_principal == '1':
        loop_chat_puro()
    elif escolha_principal == '2':
        # Verifica se o gerador principal está configurado para a nuvem
        servico_principal_config = CONFIG.get("servicos", {}).get("gerador_principal", {})
        if servico_principal_config.get("tipo") != "nuvem":
            print("\n============================== AÇÃO NECESSÁRIA ==============================")
            print("❌ ERRO: A funcionalidade 'Analisar Documentos' requer que o servidor")
            print("   esteja configurado com um modelo de NUVEM como 'gerador_principal'.")
            print(f"\n   O servidor está atualmente configurado para usar o modelo LOCAL:")
            print(f"   '{nome_modelo_principal_ativo}'")
            print("\n   Por favor, reinicie o 'servidor_modelo_local.py' e, na configuração")
            print("   do serviço 'gerador_principal', escolha a opção 'Nuvem (via OpenRouter)'.")
            print("===============================================================================")
            exit()
        loop_analise_de_arquivos()
    elif escolha_principal in opcoes_rag:
        ctx_info = opcoes_rag[escolha_principal]
        if "❌" in ctx_info["status"]:
            print(f"\nERRO: O especialista '{ctx_info['nome']}' não foi indexado.")
            exit()
        print(f"\n-> Carregando o conhecimento do '{ctx_info['nome']}'...")
        db_contexto = FAISS.load_local(os.path.join(PASTA_BASE_INDICES, ctx_info["id"]), embeddings, allow_dangerous_deserialization=True)
        loop_chat_rag(db_contexto, ctx_info["nome"])
    else:
        print("Escolha inválida.")
        
    print("\n👋 Sessão encerrada. Até logo!")